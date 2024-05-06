#import libraries 
from openpyxl import load_workbook



excelFile = 'phinCalc.xlsx'



def calculateTax(income, taxBrackets):

	taxAmount = 0.0


	previousTaxThreshold = 0.0
	for taxBracket in taxBrackets[:-1]:
		
		taxRate = taxBracket[0]/100.0
		taxThreshold = taxBracket[2]
		
		if income <= taxThreshold:
			taxableAmount = income - previousTaxThreshold
			return taxAmount + taxableAmount * taxRate
		
		taxableAmount = taxThreshold - previousTaxThreshold
		taxAmount = taxAmount + taxableAmount * taxRate

		previousTaxThreshold = taxThreshold


	taxBracket = taxBrackets[-1]

	taxRate = taxBracket[0]/100.0
	taxableAmount = income - previousTaxThreshold
	return taxAmount + taxableAmount * taxRate



def calculateFederalAndStateTax(income, federalTaxBrackets, stateTaxBrackets):

	return calculateTax(income, federalTaxBrackets) + calculateTax(income, stateTaxBrackets)



wb = load_workbook(excelFile, data_only = True)
sheet = wb.worksheets[0]


information = sheet['E7':'E12']
federalTaxRates = sheet['C17':'E23']
stateTaxRates = sheet['C28':'E36']
income = sheet['D41':'E123']
expense = sheet['F41':'G123']


additionalIncomeTax = 1.0 - information[2][0].value/100.0
capitalGainsInterest = information[3][0].value/100.0 + 1
capitalGainsTax = 1.0 - information[4][0].value/100.0
initialInvestiment = information[5][0].value
# print(additionalIncomeTax)
# print(capitalGainsInterest)
# print(capitalGainsTax)
# print(initialInvestiment)

federalTaxBrackets = []
for row in federalTaxRates:
	taxBracket = []
	for column in row:
		taxBracket.append(column.value)
	federalTaxBrackets.append(taxBracket)

stateTaxBrackets = []
for row in stateTaxRates:
	taxBracket = []
	for column in row:
		taxBracket.append(column.value)
	stateTaxBrackets.append(taxBracket)


balanceAmount = initialInvestiment
balance = []

if len(income) != len(expense):
	print('ERROR: income length does not match expense')
	quit()

for index in range(len(income)):

	balanceAmount = balanceAmount * capitalGainsInterest


	yearlyIncome = 0.0

	incomeAmount = income[index][0].value
	if incomeAmount != None:
		yearlyIncome = yearlyIncome + incomeAmount

	incomeAmount = income[index][1].value
	if incomeAmount != None:
		incomeAmount = 12 * incomeAmount
		yearlyIncome = yearlyIncome + incomeAmount

	yearlyIncome = yearlyIncome * additionalIncomeTax - calculateFederalAndStateTax(yearlyIncome, federalTaxBrackets, stateTaxBrackets)


	yearlyExpense = 0.0

	expenseAmount = expense[index][0].value
	if expenseAmount != None:
		yearlyExpense = yearlyExpense + expenseAmount

	expenseAmount = expense[index][1].value
	if expenseAmount != None:
		expenseAmount = 12 * expenseAmount
		yearlyExpense = yearlyExpense + expenseAmount


	yearlyNet = yearlyIncome - yearlyExpense
	
	if yearlyNet < 0.0:
		yearlyNet = yearlyNet / capitalGainsTax
	
	balanceAmount = balanceAmount + yearlyNet


	balance.append([balanceAmount, yearlyNet])


wb.close()



wb = load_workbook(excelFile)
sheet = wb.worksheets[0]


for rowIndex, row in enumerate(balance):
	for columnIndex, column in enumerate(row):
		sheet.cell(row = 41 + rowIndex, column = 8 + columnIndex).value = column
	print(row)


wb.save(excelFile)



